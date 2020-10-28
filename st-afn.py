import typing
from typing import Tuple
import json
import os
import argparse
import openpyxl
import math
import copy

import torch
from torch import nn
from torch import optim
from sklearn.preprocessing import StandardScaler
from sklearn.externals import joblib

import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import datetime

import utils
from modules import Encoder, Decoder
from custom_types import TCHA_Net, TrainData, TrainConfig, device
from utils import numpy_to_tvar

from torch.optim.lr_scheduler import ReduceLROnPlateau
from torch.autograd import Variable

import random


random.seed(777777)
logger = utils.setup_log()


def save_final(isMean: bool, preddata, truedata, tarRoad, timestep, interval, scaler, filename=""):
    # not change origin true data
    true = copy.deepcopy(truedata)
    pred = copy.deepcopy(preddata)

    if isMean:
        pred = pred.squeeze()
        true = true.squeeze()
        for i in range(len(pred)):
            pred[i] += scaler[-1]
            true[i] += scaler[-1]
    else:
        sca, mean = scaler.scale_[-1], scaler.mean_[-1]
        pred *= sca
        pred += mean
        true *= sca
        true += mean
        pred = pred.squeeze()
        true = true.squeeze()

    final_res = [pred, true]
    headers = ['pred', 'true', 'mae', 'mse', 'rmse', 'mape']
    wb1 = openpyxl.Workbook()
    if filename == "":
        filename = 'result/' + tarRoad + '/attn_lags' + str(timestep) + 'intervals' + str(interval) + 'ture&pred.xlsx'
    else:
        filename = filename + 'ture&pred.xlsx'
    wb1.save(filename)
    savetoexcel(final_res, headers, "sheet1", filename)

    plt.close()


def savetoexcel(data, fields, sheetname, wbname):
    wb = openpyxl.load_workbook(filename=wbname)

    sheet = wb.active
    sheet.title = sheetname

    for field in range(1, len(fields) + 1):  # 写入表头
        _ = sheet.cell(row=1, column=field, value=str(fields[field - 1]))

    for row1 in range(2, len(data[0]) + 2):  # 写入数据
        pd = data[0][row1 - 2]
        gt = data[1][row1 - 2]

        _ = sheet.cell(row=row1, column=1, value=pd)
        _ = sheet.cell(row=row1, column=2, value=gt)

        _ = sheet.cell(row=row1, column=3, value=math.fabs(pd - gt))
        _ = sheet.cell(row=row1, column=4, value=math.fabs(pd - gt) * math.fabs(pd - gt))
        _ = sheet.cell(row=row1, column=5, value=math.fabs(pd - gt) * math.fabs(pd - gt))
        _ = sheet.cell(row=row1, column=6, value=math.fabs(pd - gt) / gt)

    wb.save(filename=wbname)
    print(wbname + " save success")


def preprocess_data(dat, speed_data, col_names, mean_stand=True) -> Tuple[TrainData, StandardScaler]:
    if mean_stand:
        scale = dat.mean()
        scale = np.array(scale)
        scale_speed = speed_data.mean()
        scale_speed = np.array(scale_speed)
        proc_dat = dat
        proc_dat_speed = speed_data
        i = 0
        for col in proc_dat.columns:
            proc_dat[col] -= scale[i]
            i += 1
        proc_dat = np.array(proc_dat)

        i = 0
        for col in proc_dat_speed.columns:
            proc_dat_speed[col] -= scale_speed[i]
            i += 1
        proc_dat_speed = np.array(proc_dat_speed)
    else:
        scale = StandardScaler().fit(dat)
        proc_dat = scale.transform(dat)
        scale_speed = StandardScaler().fit(speed_data)
        proc_dat_speed = scale_speed.transform(speed_data)

    # origin data
    # proc_dat = np.array(dat)

    mask = np.ones(proc_dat.shape[1], dtype=bool)
    dat_cols = list(dat.columns)
    for col_name in col_names:
        mask[dat_cols.index(col_name)] = False

    feats = proc_dat[:, mask]
    targs = proc_dat[:, ~mask]

    return TrainData(feats, targs, proc_dat_speed), scale, scale_speed


def TCHA(train_data: TrainData, n_targs: int, bidirec=False, num_layer=1, encoder_hidden_size=64, decoder_hidden_size=64,
         T=10, learning_rate=0.01, batch_size=128, interval=1, split=0.7, isMean=False):
    train_cfg = TrainConfig(T, int(train_data.feats.shape[0] * split), batch_size, nn.MSELoss(), interval, T, isMean)
    logger.info(f"Training size: {train_cfg.train_size:d}.")

    enc_args = {"input_size": train_data.feats.shape[1], "hidden_size": encoder_hidden_size, "T": T,
                  "bidirec": bidirec, "num_layer": num_layer}
    encoder = Encoder(**enc_args).to(device)

    dec_args = {"encoder_hidden_size": encoder_hidden_size, "decoder_hidden_size": decoder_hidden_size, "T": T,
                  "out_feats": n_targs, "bidirec": bidirec, "num_layer": num_layer}
    decoder = Decoder(**dec_args).to(device)

    encoder_optimizer = optim.Adam(
        params=[p for p in encoder.parameters() if p.requires_grad],
        lr=learning_rate)
    decoder_optimizer = optim.Adam(
        params=[p for p in decoder.parameters() if p.requires_grad],
        lr=learning_rate)
    tcha = TCHA_Net(encoder, decoder, encoder_optimizer, decoder_optimizer)

    return train_cfg, tcha


def myMSE(preddata, truedata, isMean):

    true = copy.deepcopy(truedata)
    pred = copy.deepcopy(preddata)

    if isMean:
        pred = pred.squeeze()
        true = true.squeeze()
        for i in range(len(pred)):
            pred[i] += scaler[-1]
            true[i] += scaler[-1]
    else:
        sca, mean = scaler.scale_[-1], scaler.mean_[-1]
        pred *= sca
        pred += mean
        true *= sca
        true += mean
        pred = pred.squeeze()
        true = true.squeeze()

    res = 0
    for i in range(len(pred)):
        res += ((pred[i] - true[i]) ** 2)
    return res / len(pred)


def train(net: TCHA, train_data: TrainData, t_cfg: TrainConfig, tarRoad, scaler, n_epochs=10, save_plots=False):
    # 每epoch中含有批次数
    iter_per_epoch = int(np.ceil(t_cfg.train_size * 1. / t_cfg.batch_size))
    # 批次总数
    iter_losses = np.zeros(n_epochs * iter_per_epoch)
    epoch_losses = np.zeros(n_epochs)
    logger.info(f"Iterations per epoch: {t_cfg.train_size * 1. / t_cfg.batch_size:3.6f} ~ {iter_per_epoch:d}.")

    n_iter = 0

    enschedual = ReduceLROnPlateau(net.enc_opt, 'min', patience=5, factor=0.5)
    deschedual = ReduceLROnPlateau(net.dec_opt, 'min', patience=5, factor=0.5)
    # vis = visdom.Visdom()
    # winstr = 'HC train loss B: ' + str(t_cfg.batch_size) + 'I: ' + str(t_cfg.interval)
    # vis.line([0.], [0.], win=winstr, opts=dict(title=winstr))
    # global_step = 0

    for e_i in range(n_epochs):
        st = datetime.datetime.now()

        perm_idx = np.random.permutation(t_cfg.train_size - t_cfg.T - t_cfg.interval)
        # perm_idx = np.arange(0, t_cfg.train_size - t_cfg.T - t_cfg.interval)

        for t_i in range(0, t_cfg.train_size - t_cfg.T - t_cfg.interval, t_cfg.batch_size):
            # 每一混乱批次
            # if t_i == 5632:
            #     print(t_i)
            endidx = min(t_i + t_cfg.batch_size, t_cfg.train_size)
            batch_idx = perm_idx[t_i:endidx]
            feats, y_history, y_target, speed = prep_train_data(batch_idx, t_cfg, train_data)

            loss = train_iteration(net, t_cfg.loss_func, feats, y_history, y_target, speed)

            # 第t_i批次的loss放入
            iter_losses[e_i * iter_per_epoch + t_i // t_cfg.batch_size] = loss
            # if (j / t_cfg.batch_size) % 50 == 0:
            #    self.logger.info("Epoch %d, Batch %d: loss = %3.3f.", i, j / t_cfg.batch_size, loss)
            n_iter += 1
            # adjust_learning_rate(net, n_iter)
        # 当前epoch平均loss
        epoch_losses[e_i] = np.mean(iter_losses[range(e_i * iter_per_epoch, (e_i + 1) * iter_per_epoch)])
        # vis.line([epoch_losses[e_i]], [global_step], win=winstr, update='append')
        # global_step += 1
        enschedual.step(epoch_losses[e_i])
        deschedual.step(epoch_losses[e_i])
        # if e_i % 20 == 0:
        weights, y_test_pred, _ = predict(net, train_data,
                                 t_cfg.train_size, t_cfg.batch_size, t_cfg.T, t_cfg.interval, on_train=False)
        # TODO: make this MSE and make it work for multiple inputs
        # val_loss = y_test_pred - train_data.targs[t_cfg.train_size:]
        val_loss = myMSE(y_test_pred, train_data.targs[t_cfg.train_size:], t_cfg.isMean)

        save_final(t_cfg.isMean, y_test_pred, train_data.targs[t_cfg.train_size:], tarRoad, config.timestep, config.interval, scaler)

        weights, y_train_pred, _ = predict(net, train_data,
                                  t_cfg.train_size, t_cfg.batch_size, t_cfg.T, t_cfg.interval, on_train=True)

        save_final(t_cfg.isMean, y_train_pred, train_data.targs[:t_cfg.train_size], tarRoad, config.timestep, config.interval, scaler,filename="hhh")

        train_loss = myMSE(y_train_pred, train_data.targs[:t_cfg.train_size], t_cfg.isMean)

        logger.info(f"Epoch {e_i:d}, train loss: {np.mean(np.abs(epoch_losses[e_i]))}, val loss: {np.mean(np.abs(val_loss))}.")

        torch.save(net.encoder.state_dict(),
                   'models/' + tarRoad + '/' + str(t_cfg.interval) + '/HCAdam_encoder' + str(t_cfg.batch_size) +
                   str(e_i) + '-norm' + '.model')
        torch.save(net.decoder.state_dict(),
                   'models/' + tarRoad + '/' + str(t_cfg.interval) + '/HCAdam_decoder' + str(t_cfg.batch_size) +
                   str(e_i) + '-norm' + '.model')
        plt.figure()
        plt.plot(range(1, 1 + len(train_data.targs)), train_data.targs,
                 label="True")
        plt.plot(range(t_cfg.T, len(y_train_pred) + t_cfg.T), y_train_pred,
                 label='Predicted - Train')
        plt.plot(range(t_cfg.T + len(y_train_pred) + t_cfg.interval, len(train_data.targs)), y_test_pred,
                 label='Predicted - Test')
        plt.legend(loc='upper left')
        savename = "pred_epoch" + str(e_i) + "interval" + str(t_cfg.interval) + 'batchsize' + str(
            t_cfg.batch_size) + '.png'
        utils.save_or_show_plot(savename, save_plots, t_cfg.interval, tarRoad)
        pd_iterloss = pd.DataFrame(iter_losses)
        pd_epoloss = pd.DataFrame(epoch_losses)
        pd_iterloss.to_csv('result/' + tarRoad + '/iterloss iterval' + str(t_cfg.interval) + '.csv')
        pd_epoloss.to_csv('result/' + tarRoad + '/epochloss iterval' + str(t_cfg.interval) + '.csv')
        plt.close()

        ed = datetime.datetime.now()
        print('epoch:{}, time cost:{}\n\n'.format(e_i, (ed - st).seconds))

    return iter_losses, epoch_losses


def prep_train_data(batch_idx: np.ndarray, t_cfg: TrainConfig, train_data: TrainData):
    feats = np.zeros((len(batch_idx), t_cfg.T, train_data.feats.shape[1]))
    y_history = np.zeros((len(batch_idx), t_cfg.T, train_data.targs.shape[1]))
    speed = np.zeros((len(batch_idx), t_cfg.T, train_data.speeds.shape[1]))
    y_target = train_data.targs[batch_idx + t_cfg.T + t_cfg.interval - 1]

    for b_i, b_idx in enumerate(batch_idx):
        b_slc = slice(b_idx, b_idx + t_cfg.T)
        feats[b_i, :, :] = train_data.feats[b_slc, :]
        y_history[b_i, :] = train_data.targs[b_slc]
        speed[b_i, :] = train_data.speeds[b_slc]
    return feats, y_history, y_target, speed


def adjust_learning_rate(net: TCHA, n_iter: int):
    # TODO: Where did this Learning Rate adjustment schedule come from?
    # Should be modified to use Cosine Annealing with warm restarts https://www.jeremyjordan.me/nn-learning-rate/
    if n_iter % 100 == 0 and n_iter > 0:
        for enc_params, dec_params in zip(net.enc_opt.param_groups, net.dec_opt.param_groups):
            enc_params['lr'] = enc_params['lr'] * 0.9
            dec_params['lr'] = dec_params['lr'] * 0.9


def train_iteration(t_net: TCHA, loss_func: typing.Callable, X, y_history, y_target, speed):


    X = numpy_to_tvar(X)
    input_weighted, input_encoded = t_net.encoder(X)
    _,y_pred = t_net.decoder(input_encoded, numpy_to_tvar(y_history), numpy_to_tvar(speed))
    y_true = numpy_to_tvar(y_target)

    loss = loss_func(y_pred, y_true)
    # loss = loss_func(numpy_to_tvar(y_pred[1]), y_true)

    t_net.enc_opt.zero_grad()
    t_net.dec_opt.zero_grad()
    loss.backward()

    t_net.enc_opt.step()
    t_net.dec_opt.step()


    return loss.item()


def predict(t_net: TCHA, t_dat: TrainData, train_size: int, batch_size: int, T: int, interval: int, on_train=False):

    # summary_temporal = np.zeros(T)
    # summary_spatial = np.zeros((T, t_dat.feats.shape[1]))
    summary_temporal, summary_spatial = [], []
    embeded_weights = np.zeros((T, t_dat.feats.shape[1]))
    count = 0
    out_size = t_dat.targs.shape[1]
    if on_train:
        y_pred = np.zeros((train_size - T - interval, out_size))
        y_tar = np.zeros((train_size - T - interval, out_size))
    else:
        y_pred = np.zeros((t_dat.feats.shape[0] - train_size, out_size))
        y_tar = np.zeros((t_dat.feats.shape[0] - train_size, out_size))

    p = 0

    for y_i in range(0, len(y_pred) - interval, batch_size):
        y_slc = slice(y_i, y_i + batch_size)
        # batch_idx：批次开始、结束的下标
        batch_idx = range(len(y_pred))[y_slc]
        b_len = len(batch_idx)
        X = np.zeros((b_len, T, t_dat.feats.shape[1]))
        y_history = np.zeros((b_len, T, t_dat.targs.shape[1]))
        speed = np.zeros((b_len, T, t_dat.speeds.shape[1]))

        for b_i, b_idx in enumerate(batch_idx):
            # b_i：每批次下标号，需清0
            # b_idx：开始下标下标号码，不清0
            if on_train:
                idx = range(b_idx, b_idx + T)
            else:
                idx = range(b_idx + train_size - T - interval, b_idx + train_size - interval)

            X[b_i, :, :] = t_dat.feats[idx, :]
            y_history[b_i, :] = t_dat.targs[idx]
            speed[b_i, :] = t_dat.speeds[idx]

            y_tar[p] = t_dat.targs[idx[-1] + interval]
            p += 1

        y_history = numpy_to_tvar(y_history)
        speed = numpy_to_tvar(speed)
        spatial_weights, input_encoded = t_net.encoder(numpy_to_tvar(X))
        temporal_weights, pred = t_net.decoder(input_encoded, y_history, speed)
        y_pred[y_slc] = pred.cpu().data.numpy()

        # following comment: weights to be drawn
        spatial_weights = spatial_weights.cpu().detach().numpy()
        temporal_weights = temporal_weights.cpu().detach().numpy()
        count += spatial_weights.shape[0]
        # summary_spatial += spatial_weights.mean(axis=0)
        # summary_temporal += temporal_weights.mean(axis=0)
        for cnt in range(temporal_weights.shape[0]):
            # 每一批次中各路段平均
            summary_spatial.append(spatial_weights.mean(axis=1)[cnt])
            # 每一批次中各时段平均
            summary_temporal.append(temporal_weights[cnt])
        ss = np.array(summary_spatial)
        st = np.array(summary_temporal)

    # summary_temporal /= count
    # summary_spatial /= count
    # print(summary_spatial)
    # for i in range(summary_spatial.shape[1]):
    #     for j in range(T):
    #         embeded_weights[j][i] = summary_spatial[j][i] + summary_temporal[T - 1 - j]
    # pdSS, pdST = pd.DataFrame(ss), pd.DataFrame(st)
    # pdSS.to_excel('spatial weights.xls')
    # pdST.to_excel('temporal weights.xls')
    return (ss, st), y_pred, y_tar


def getArgParser():
    parser = argparse.ArgumentParser(description='TCHA for predicting traffic speed')
    parser.add_argument(
        '-e', '--epoch', type=int, default=40,
        help='the number of epochs')
    parser.add_argument(
        '-b', '--batch_size', type=int, default=128,
        help='the mini-batch size')
    parser.add_argument(
        '-s', '--split', type=float, default=0.7,
        help='the split ratio of validation set')
    parser.add_argument(
        '-i', '--intervals', type=int, default=1,
        help='save models every interval epoch')
    parser.add_argument(
        '-lr', '--lrate', type=float, default=1e-4,
        help='learning rate')
    parser.add_argument(
        '-t', '--test', action='store_true', default=False,
        help='train or test')
    parser.add_argument(
        '-n', '--number', type=str, default='39',
        help='model number')
    parser.add_argument(
        '-g', '--time_step', type=int, default=12
    )
    parser.add_argument(
        '-m', '--mean_stand', type=bool, default=True
    )
    parser.add_argument(
        '-r', '--road', type=str, default='tonghui'
    )
    parser.add_argument(
        '-bi', '--bidirec', type=bool, default=True
    )
    parser.add_argument(
        '-num', '--numLayer', type=int, default=2
    )
    return parser


if __name__ == '__main__':
    args = getArgParser().parse_args()
    num_epochs = args.epoch
    batch_size = args.batch_size
    split = args.split
    # interval = args.intervals
    timestep = args.time_step
    lr = args.lrate
    test = args.test
    modelNo = args.number
    save_plots = True
    debug = False
    isMean = args.mean_stand
    tarRoad = args.road
    bidirec = args.bidirec
    numLayer = args.numLayer

    interval = 1
    while interval < 2:

        # target road(column) name
        tarcol = "n-z_3" if tarRoad == 'tonghui' else 'boxue_jinhui'

        raw_data = pd.read_csv(os.path.join("data/" + tarRoad, "all_add.csv"), nrows=100 if debug else None)
        if tarcol == "changyuan_2":
            raw_speed = pd.read_csv(os.path.join("data/" + tarRoad, "speed_lane2.csv"), nrows=100 if debug else None)
        else:
            raw_speed = pd.read_csv(os.path.join("data/" + tarRoad, "speed_lane3.csv"), nrows=100 if debug else None)
        logger.info(f"num of epoches: {num_epochs}. batch size: {batch_size}. lr: {lr}. \n")
        logger.info(f"Shape of data: {raw_data.shape}.\nMissing in data: {raw_data.isnull().sum().sum()}.")
        targ_cols = (tarcol,)
        data, scaler, scaler_speed = preprocess_data(raw_data, raw_speed, targ_cols, isMean)



        tcha_args = {"batch_size": batch_size, "T": timestep, "interval": interval, "split": split,
                         "bidirec": bidirec, "num_layer": numLayer, "isMean": isMean}
        config, model = TCHA(data, n_targs=len(targ_cols), learning_rate=lr, **tcha_args)
        if test:
            encoder_name = 'models/' + tarRoad + '/' + str(interval) + '/HCAdam_encoder' + str(batch_size) + modelNo + '-norm.model'
            decoder_name = 'models/' + tarRoad + '/' + str(interval) + '/HCAdam_decoder' + str(batch_size) + modelNo + '-norm.model'
            model.encoder.load_state_dict(torch.load(encoder_name, map_location=lambda storage, loc: storage))
            model.decoder.load_state_dict(torch.load(decoder_name, map_location=lambda storage, loc: storage))
            embeded_weights, final_y_pred, y_true = predict(model, data, config.train_size, config.batch_size, config.T, config.interval)
            pd_sptial = pd.DataFrame(embeded_weights[0])
            pd_time = pd.DataFrame(embeded_weights[1])
            pd_sptial.to_csv('result/' + 'att_sptial'+'.csv')
            pd_time.to_csv('result/' + 'att_time' + '.csv')

        else:
            iter_loss, epoch_loss = train(model, data, config, tarRoad, scaler, n_epochs=num_epochs, save_plots=save_plots)
            weightss, final_y_pred, y_true = predict(model, data, config.train_size, config.batch_size, config.T, config.interval)
            pd_sptial = pd.DataFrame(weightss[0])
            pd_time = pd.DataFrame(weightss[1])
            pd_sptial.to_csv('result/' + 'att_sptial'+'.csv')
            pd_time.to_csv('result/' + 'att_time' + '.csv')
            try:
                plt.figure()
                plt.semilogy(range(len(iter_loss)), iter_loss)
                utils.save_or_show_plot("train iter_loss interval" + str(interval) + ".png", save_plots, interval=interval, tarRoad=tarRoad)

                plt.figure()
                plt.semilogy(range(len(epoch_loss)), epoch_loss)
                utils.save_or_show_plot("train epoch epoch_loss interval" + str(interval) + ".png", save_plots, interval, tarRoad)
            except Exception as e:
                print(e)

        plt.figure()
        plt.plot(data.targs[config.train_size:], label="True")
        plt.plot(final_y_pred, label='Predicted')
        plt.legend(loc='upper left')
        # y1 = pd.DataFrame(y_true)
        # y2 = pd.DataFrame(data.targs[config.train_size:])
        # print(y1)
        # print(y2)
        filename = "shixinTCHAfinal_predictedbatchsize" + str(batch_size) + "interval" + str(interval)
        utils.save_or_show_plot(filename + ".png", save_plots, interval, tarRoad=tarRoad)

        save_final(isMean, final_y_pred, y_true, tarRoad, timestep, interval, scaler, filename)
        interval += 1
